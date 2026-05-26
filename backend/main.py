"""角色资产库后端服务。

启动： python main.py     （或 uvicorn main:app --reload --port 8000）
前端使用同源 /api 与 /images 路径访问后端。
"""
import io
import os
import re
import zipfile
from pathlib import Path
from time import monotonic
from typing import Callable, Optional
from urllib.parse import quote

from dotenv import load_dotenv

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(BASE_DIR, ".env"))  # 仅加载本项目的 .env

for _p in ("HTTP_PROXY", "HTTPS_PROXY", "ALL_PROXY",
           "http_proxy", "https_proxy", "all_proxy"):
    os.environ.pop(_p, None)

from fastapi import FastAPI, File, HTTPException, Query, Response, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import RedirectResponse
from pydantic import BaseModel, field_validator

import ai
import storage
from db import (
    ASSET_CHARACTER,
    ASSET_SCENE,
    CHARACTER_FIELDS,
    FILTER_FIELDS,
    SCENE_FIELDS,
    SCENE_FILTER_FIELDS,
    asset_to_dict,
    attach_image,
    create_asset,
    create_media_asset,
    create_video_benchmark_item,
    delete_asset,
    delete_image as db_delete_image,
    delete_video_benchmark_item,
    get_asset,
    get_conn,
    get_options as db_get_options,
    get_video_benchmark_item,
    health_check as db_health_check,
    init_db,
    list_assets,
    list_media_assets as db_list_media_assets,
    list_video_benchmark_items,
    now,
    replace_source_images,
    set_cover as db_set_cover,
    update_asset,
    update_video_benchmark_item,
)

app = FastAPI(title="角色与场景资产库")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # 内部工具，访问控制交给 Nginx Basic Auth
    allow_methods=["*"],
    allow_headers=["*"],
)

init_db()

API_CACHE_TTL_SECONDS = float(os.getenv("API_CACHE_TTL_SECONDS", "30"))
_API_CACHE: dict[tuple, tuple[float, object]] = {}


def _cached_api(key: tuple, loader: Callable[[], object]) -> object:
    if API_CACHE_TTL_SECONDS <= 0:
        return loader()
    now_ts = monotonic()
    cached = _API_CACHE.get(key)
    if cached and now_ts - cached[0] < API_CACHE_TTL_SECONDS:
        return cached[1]
    value = loader()
    _API_CACHE[key] = (now_ts, value)
    return value


def _clear_api_cache() -> None:
    _API_CACHE.clear()


class CharacterIn(BaseModel):
    era: str = ""
    type: str = ""
    gender: str = ""
    age: str = ""
    persona: str = ""
    body: str = ""
    features: str = ""
    genre: str = ""
    prompt: str = ""
    description: str = ""


class PromptReq(CharacterIn):
    pass


class ImageGenReq(BaseModel):
    prompt: str
    set_cover: bool = False


class ExtractReq(BaseModel):
    description: str = ""


class SceneIn(BaseModel):
    name: str = ""
    era: str = ""
    scene_type: str = ""
    genre: str = ""
    mood: str = ""
    elements: str = ""
    prompt: str = ""
    description: str = ""


class ScenePromptReq(SceneIn):
    pass


class SceneViewReq(BaseModel):
    view: str  # reverse | multiview


class VideoBenchmarkItemIn(BaseModel):
    shot_type: str = ""
    task_type: str = ""
    question_type: str = ""
    scene: str = ""
    screen_size: str = ""
    character_image_asset: str = ""
    scene_image_asset: str = ""
    prop_image_asset: str = ""
    audio_input: str = ""
    video_input: str = ""
    text_prompt: str = ""
    judging_criteria: str = ""
    video_output: str = ""
    score: Optional[int] = None
    character_image_id: Optional[int] = None
    scene_image_id: Optional[int] = None
    prop_image_id: Optional[int] = None
    audio_input_id: Optional[int] = None
    video_input_id: Optional[int] = None
    video_output_id: Optional[int] = None
    character_image_ids: list[int] = []
    scene_image_ids: list[int] = []
    prop_image_ids: list[int] = []
    audio_input_media_ids: list[int] = []
    video_input_ids: list[int] = []
    video_output_ids: list[int] = []

    @field_validator("score")
    @classmethod
    def validate_score(cls, value: Optional[int]) -> Optional[int]:
        if value is not None and not 0 <= value <= 5:
            raise ValueError("score must be between 0 and 5")
        return value


VIEW_PROMPTS = {
    "reverse": (
        "The attached image is a reference of a scene environment. Create ONE "
        "wide image divided into 2 equal panels showing this EXACT SAME "
        "environment from two opposite camera directions (shot / reverse-shot): "
        "LEFT panel from one direction, RIGHT panel the reverse view with the "
        "camera turned 180 degrees. Keep architecture, layout, objects, "
        "materials, colors and lighting fully consistent with the reference. "
        "No people. Photorealistic, cinematic lighting, ultra-detailed.",
        "16:9",
    ),
    "multiview": (
        "The attached image is a reference of a scene environment. Create ONE "
        "image arranged as a 2x2 grid of 4 panels showing this EXACT SAME "
        "environment from 4 DISTINCTLY DIFFERENT camera angles, as if the camera "
        "rotates 90 degrees between each panel to face a different direction of "
        "the space: front, right side, the opposite (rear) direction, and left "
        "side. The four panels MUST be clearly different viewpoints, not minor "
        "variations of the same shot. Keep architecture, layout, objects, "
        "materials, colors and lighting consistent with the reference. No "
        "people. Photorealistic, cinematic lighting, ultra-detailed.",
        "1:1",
    ),
}
VIEW_SOURCES = set(VIEW_PROMPTS)


def fetch_character(conn, cid: int) -> dict:
    row = get_asset(conn, ASSET_CHARACTER, cid)
    if row is None:
        raise HTTPException(404, "角色不存在")
    return asset_to_dict(conn, row, CHARACTER_FIELDS)


def fetch_scene(conn, sid: int) -> dict:
    row = get_asset(conn, ASSET_SCENE, sid)
    if row is None:
        raise HTTPException(404, "场景不存在")
    return asset_to_dict(conn, row, SCENE_FIELDS, VIEW_SOURCES)


def fetch_video_benchmark_item(conn, item_id: int) -> dict:
    item = get_video_benchmark_item(conn, item_id)
    if item is None:
        raise HTTPException(404, "视频 Benchmark 记录不存在")
    return item


def _delete_objects(keys: list[str]) -> None:
    for key in keys:
        try:
            storage.delete_object(key)
        except Exception:
            pass


def _upload_image_bytes(data: bytes, ext: str = ".png", content_type: str = "image/png") -> str:
    key = storage.new_object_key(ext)
    storage.put_bytes(key, data, content_type)
    return key


def _upload_media_bytes(data: bytes, filename: str, content_type: str, media_type: str) -> str:
    ext = Path(filename or "").suffix or (".mp4" if media_type == "video" else ".mp3" if media_type == "audio" else ".png")
    key = storage.new_object_key(ext, prefix=f"{media_type}s")
    storage.put_bytes(key, data, content_type or "application/octet-stream")
    return key


@app.get("/images/{object_key:path}")
def image_redirect(object_key: str):
    """Keep the old /images/... URL shape while serving private TOS objects."""
    try:
        return RedirectResponse(storage.object_url(object_key))
    except Exception as e:  # noqa: BLE001
        raise HTTPException(404, f"图片不可用：{e}") from e


@app.get("/api/video-benchmark-items")
def list_video_benchmark_api(
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    q: Optional[str] = None,
    shot_type: Optional[str] = None,
    task_type: Optional[str] = None,
    question_type: Optional[str] = None,
    scene: Optional[str] = None,
    screen_size: Optional[str] = None,
    score: Optional[int] = Query(None, ge=0, le=5),
):
    with get_conn() as conn:
        return list_video_benchmark_items(
            conn,
            limit=limit,
            offset=offset,
            q=q,
            shot_type=shot_type,
            task_type=task_type,
            question_type=question_type,
            scene=scene,
            screen_size=screen_size,
            score=score,
        )


@app.get("/api/media-assets")
def list_media_assets_api(
    media_type: Optional[str] = Query(None, pattern="^(image|audio|video)$"),
    asset_kind: Optional[str] = Query(None, pattern="^(character|scene|audio|prop|video)$"),
    q: Optional[str] = None,
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
):
    with get_conn() as conn:
        return db_list_media_assets(
            conn,
            media_type=media_type,
            asset_kind=asset_kind,
            q=q,
            limit=limit,
            offset=offset,
        )


@app.post("/api/media-assets/upload")
async def upload_media_asset_api(
    file: UploadFile = File(...),
    media_type: str = Query(..., pattern="^(image|audio|video)$"),
    asset_kind: str = Query(..., pattern="^(character|scene|audio|prop|video)$"),
    title: Optional[str] = None,
):
    if media_type == "image" and not (file.content_type or "").startswith("image/"):
        raise HTTPException(422, "图片素材只能上传 image/* 文件")
    if media_type == "audio" and not (file.content_type or "").startswith("audio/"):
        raise HTTPException(422, "音频素材只能上传 audio/* 文件")
    if media_type == "video" and not (file.content_type or "").startswith("video/"):
        raise HTTPException(422, "视频素材只能上传 video/* 文件")
    try:
        key = _upload_media_bytes(
            await file.read(),
            file.filename or "",
            file.content_type or "",
            media_type,
        )
        with get_conn() as conn:
            media = create_media_asset(
                conn,
                asset_kind,
                media_type,
                key,
                title or file.filename or key,
            )
            conn.commit()
            return media
    except ValueError as e:
        raise HTTPException(422, str(e)) from e


@app.get("/api/video-benchmark-items/{item_id}")
def get_video_benchmark_api(item_id: int):
    with get_conn() as conn:
        return fetch_video_benchmark_item(conn, item_id)


@app.post("/api/video-benchmark-items")
def create_video_benchmark_api(payload: VideoBenchmarkItemIn):
    with get_conn() as conn:
        try:
            item_id = create_video_benchmark_item(conn, payload)
        except ValueError as e:
            raise HTTPException(422, str(e)) from e
        conn.commit()
        return fetch_video_benchmark_item(conn, item_id)


@app.put("/api/video-benchmark-items/{item_id}")
def update_video_benchmark_api(item_id: int, payload: VideoBenchmarkItemIn):
    with get_conn() as conn:
        try:
            if not update_video_benchmark_item(conn, item_id, payload):
                raise HTTPException(404, "视频 Benchmark 记录不存在")
        except ValueError as e:
            raise HTTPException(422, str(e)) from e
        conn.commit()
        return fetch_video_benchmark_item(conn, item_id)


@app.delete("/api/video-benchmark-items/{item_id}")
def delete_video_benchmark_api(item_id: int):
    with get_conn() as conn:
        if not delete_video_benchmark_item(conn, item_id):
            raise HTTPException(404, "视频 Benchmark 记录不存在")
        conn.commit()
    return {"ok": True}


@app.get("/api/options")
def get_options():
    def load():
        with get_conn() as conn:
            return db_get_options(conn, ASSET_CHARACTER, FILTER_FIELDS)

    return _cached_api(("character_options",), load)


@app.get("/api/characters")
def list_characters(
    era: Optional[str] = None,
    type: Optional[str] = None,
    gender: Optional[str] = None,
    age: Optional[str] = None,
    genre: Optional[str] = None,
    q: Optional[str] = None,
):
    def load():
        with get_conn() as conn:
            result = list_assets(
                conn,
                ASSET_CHARACTER,
                CHARACTER_FIELDS,
                [("era", era), ("type", type), ("gender", gender), ("age", age), ("genre", genre)],
                q,
                ["persona", "features", "prompt"],
            )
        result.sort(key=lambda x: (x.get("genre") or "", x["id"]))
        return result

    return _cached_api(("characters", era, type, gender, age, genre, q), load)


@app.get("/api/characters/{cid}")
def get_character(cid: int):
    with get_conn() as conn:
        return fetch_character(conn, cid)


@app.post("/api/characters")
def create_character(payload: CharacterIn):
    with get_conn() as conn:
        cid = create_asset(conn, ASSET_CHARACTER, CHARACTER_FIELDS, payload)
        conn.commit()
        data = fetch_character(conn, cid)
    _clear_api_cache()
    return data


@app.put("/api/characters/{cid}")
def update_character(cid: int, payload: CharacterIn):
    with get_conn() as conn:
        fetch_character(conn, cid)
        update_asset(conn, ASSET_CHARACTER, cid, CHARACTER_FIELDS, payload)
        conn.commit()
        data = fetch_character(conn, cid)
    _clear_api_cache()
    return data


@app.delete("/api/characters/{cid}")
def delete_character(cid: int):
    with get_conn() as conn:
        fetch_character(conn, cid)
        keys = delete_asset(conn, ASSET_CHARACTER, cid)
        conn.commit()
    _clear_api_cache()
    _delete_objects(keys)
    return {"ok": True}


@app.post("/api/characters/{cid}/images")
async def upload_image(cid: int, file: UploadFile = File(...)):
    with get_conn() as conn:
        fetch_character(conn, cid)
        ext = os.path.splitext(file.filename or "")[1].lower() or ".png"
        content_type = file.content_type or "application/octet-stream"
        key = _upload_image_bytes(await file.read(), ext, content_type)
        img = attach_image(conn, cid, key, "uploaded")
        conn.commit()
    _clear_api_cache()
    return img


@app.delete("/api/images/{img_id}")
def delete_image(img_id: int):
    with get_conn() as conn:
        key = db_delete_image(conn, img_id, ASSET_CHARACTER)
        if key is None:
            raise HTTPException(404, "图片不存在")
        conn.commit()
    _clear_api_cache()
    storage.delete_object(key)
    return {"ok": True}


@app.put("/api/characters/{cid}/cover/{img_id}")
def set_cover(cid: int, img_id: int):
    with get_conn() as conn:
        fetch_character(conn, cid)
        try:
            db_set_cover(conn, ASSET_CHARACTER, cid, img_id)
        except KeyError as e:
            raise HTTPException(404, "图片不存在") from e
        conn.commit()
        data = fetch_character(conn, cid)
    _clear_api_cache()
    return data


@app.post("/api/extract-fields")
def api_extract_fields(payload: ExtractReq):
    if not payload.description.strip():
        raise HTTPException(400, "自由描述为空")
    try:
        return ai.extract_fields(payload.description, get_options())
    except ai.AIConfigError as e:
        raise HTTPException(400, str(e)) from e
    except ValueError as e:
        raise HTTPException(400, str(e)) from e
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f"字段解析失败：{e}") from e


@app.post("/api/generate-prompt")
def api_generate_prompt(payload: PromptReq):
    try:
        prompt = ai.generate_prompt(payload.model_dump(), payload.description)
    except ai.AIConfigError as e:
        raise HTTPException(400, str(e)) from e
    except ValueError as e:
        raise HTTPException(400, str(e)) from e
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f"提示词生成失败：{e}") from e
    return {"prompt": prompt}


@app.post("/api/characters/{cid}/generate-image")
def api_generate_image(cid: int, payload: ImageGenReq):
    with get_conn() as conn:
        fetch_character(conn, cid)

    try:
        raw = ai.generate_image(payload.prompt)
        key = _upload_image_bytes(raw)
    except ai.AIConfigError as e:
        raise HTTPException(400, str(e)) from e
    except ValueError as e:
        raise HTTPException(400, str(e)) from e
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f"图片生成失败：{e}") from e

    with get_conn() as conn:
        fetch_character(conn, cid)
        img = attach_image(conn, cid, key, "generated")
        if payload.set_cover:
            db_set_cover(conn, ASSET_CHARACTER, cid, img["id"])
        conn.commit()
    _clear_api_cache()
    return img


@app.get("/api/scenes/options")
def get_scene_options():
    def load():
        with get_conn() as conn:
            return db_get_options(conn, ASSET_SCENE, SCENE_FILTER_FIELDS)

    return _cached_api(("scene_options",), load)


@app.get("/api/scenes")
def list_scenes(
    era: Optional[str] = None,
    scene_type: Optional[str] = None,
    genre: Optional[str] = None,
    mood: Optional[str] = None,
    q: Optional[str] = None,
):
    def load():
        with get_conn() as conn:
            result = list_assets(
                conn,
                ASSET_SCENE,
                SCENE_FIELDS,
                [("era", era), ("scene_type", scene_type), ("genre", genre), ("mood", mood)],
                q,
                ["name", "elements", "prompt"],
                VIEW_SOURCES,
            )
        result.sort(key=lambda x: (x.get("genre") or "", x["id"]))
        return result

    return _cached_api(("scenes", era, scene_type, genre, mood, q), load)


@app.get("/api/scenes/{sid}")
def get_scene(sid: int):
    with get_conn() as conn:
        return fetch_scene(conn, sid)


@app.post("/api/scenes")
def create_scene(payload: SceneIn):
    with get_conn() as conn:
        sid = create_asset(conn, ASSET_SCENE, SCENE_FIELDS, payload)
        conn.commit()
        data = fetch_scene(conn, sid)
    _clear_api_cache()
    return data


@app.put("/api/scenes/{sid}")
def update_scene(sid: int, payload: SceneIn):
    with get_conn() as conn:
        fetch_scene(conn, sid)
        update_asset(conn, ASSET_SCENE, sid, SCENE_FIELDS, payload)
        conn.commit()
        data = fetch_scene(conn, sid)
    _clear_api_cache()
    return data


@app.delete("/api/scenes/{sid}")
def delete_scene(sid: int):
    with get_conn() as conn:
        fetch_scene(conn, sid)
        keys = delete_asset(conn, ASSET_SCENE, sid)
        conn.commit()
    _clear_api_cache()
    _delete_objects(keys)
    return {"ok": True}


@app.post("/api/scenes/{sid}/images")
async def upload_scene_image(sid: int, file: UploadFile = File(...)):
    with get_conn() as conn:
        fetch_scene(conn, sid)
        ext = os.path.splitext(file.filename or "")[1].lower() or ".png"
        content_type = file.content_type or "application/octet-stream"
        key = _upload_image_bytes(await file.read(), ext, content_type)
        img = attach_image(conn, sid, key, "uploaded")
        conn.commit()
    _clear_api_cache()
    return img


@app.delete("/api/scene-images/{img_id}")
def delete_scene_image(img_id: int):
    with get_conn() as conn:
        key = db_delete_image(conn, img_id, ASSET_SCENE)
        if key is None:
            raise HTTPException(404, "图片不存在")
        conn.commit()
    _clear_api_cache()
    storage.delete_object(key)
    return {"ok": True}


@app.put("/api/scenes/{sid}/cover/{img_id}")
def set_scene_cover(sid: int, img_id: int):
    with get_conn() as conn:
        fetch_scene(conn, sid)
        try:
            db_set_cover(conn, ASSET_SCENE, sid, img_id)
        except KeyError as e:
            raise HTTPException(404, "图片不存在") from e
        conn.commit()
        data = fetch_scene(conn, sid)
    _clear_api_cache()
    return data


@app.post("/api/scenes/generate-prompt")
def api_generate_scene_prompt(payload: ScenePromptReq):
    try:
        prompt = ai.generate_scene_prompt(payload.model_dump(), payload.description)
    except ai.AIConfigError as e:
        raise HTTPException(400, str(e)) from e
    except ValueError as e:
        raise HTTPException(400, str(e)) from e
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f"提示词生成失败：{e}") from e
    return {"prompt": prompt}


@app.post("/api/scenes/extract-fields")
def api_extract_scene_fields(payload: ExtractReq):
    if not payload.description.strip():
        raise HTTPException(400, "自由描述为空")
    try:
        return ai.extract_scene_fields(payload.description, get_scene_options())
    except ai.AIConfigError as e:
        raise HTTPException(400, str(e)) from e
    except ValueError as e:
        raise HTTPException(400, str(e)) from e
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f"字段解析失败：{e}") from e


@app.post("/api/scenes/{sid}/generate-image")
def api_generate_scene_image(sid: int, payload: ImageGenReq):
    with get_conn() as conn:
        fetch_scene(conn, sid)

    try:
        raw = ai.generate_image(payload.prompt)
        key = _upload_image_bytes(raw)
    except ai.AIConfigError as e:
        raise HTTPException(400, str(e)) from e
    except ValueError as e:
        raise HTTPException(400, str(e)) from e
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f"图片生成失败：{e}") from e

    with get_conn() as conn:
        fetch_scene(conn, sid)
        img = attach_image(conn, sid, key, "generated")
        if payload.set_cover:
            db_set_cover(conn, ASSET_SCENE, sid, img["id"])
        conn.commit()
    _clear_api_cache()
    return img


@app.post("/api/scenes/{sid}/generate-view")
def api_generate_scene_view(sid: int, payload: SceneViewReq):
    if payload.view not in VIEW_PROMPTS:
        raise HTTPException(400, "view 取值应为 reverse 或 multiview")
    with get_conn() as conn:
        scene = fetch_scene(conn, sid)
        cover = scene.get("cover_filename")
        if not cover:
            raise HTTPException(400, "该场景还没有图片，请先生成场景图")

    prompt, aspect = VIEW_PROMPTS[payload.view]
    try:
        ref = storage.get_bytes(cover)
        raw = ai.generate_image(prompt, ref_image_bytes=ref, aspect_override=aspect)
        key = _upload_image_bytes(raw)
    except ai.AIConfigError as e:
        raise HTTPException(400, str(e)) from e
    except ValueError as e:
        raise HTTPException(400, str(e)) from e
    except Exception as e:  # noqa: BLE001
        raise HTTPException(502, f"图片生成失败：{e}") from e

    with get_conn() as conn:
        fetch_scene(conn, sid)
        old_keys = replace_source_images(conn, sid, payload.view)
        img = attach_image(conn, sid, key, payload.view)
        conn.commit()
    _clear_api_cache()
    _delete_objects(old_keys)
    return img


CHAR_EXPORT_COLS = [
    ("era", "时代"), ("type", "类型"), ("gender", "性别"), ("age", "年龄段"),
    ("persona", "人设"), ("body", "身材"), ("features", "特征"),
    ("genre", "常见题材"), ("prompt", "人物生成提示词"),
    ("__imgname__", "原图文件名"), ("__image__", "图片"),
]
SCENE_EXPORT_COLS = [
    ("name", "场景名称"), ("era", "时代"), ("scene_type", "场景类型"),
    ("genre", "题材风格"), ("mood", "氛围时段"), ("elements", "关键元素"),
    ("prompt", "场景生成提示词"),
    ("__imgname__", "原图文件名"), ("__image__", "图片"),
]


def _compress_image(data: bytes):
    from PIL import Image as PILImage
    from openpyxl.drawing.image import Image as XLImage

    im = PILImage.open(io.BytesIO(data))
    if im.mode != "RGB":
        im = im.convert("RGB")
    buf = io.BytesIO()
    im.save(buf, format="JPEG", quality=85)
    buf.seek(0)
    return XLImage(buf)


def _sanitize(name: str) -> str:
    s = re.sub(r'[/\\:*?"<>|\x00-\x1f]', "_", (name or "").strip())
    return (s or "未命名")[:60]


def _export_filenames(rows: list, name_key: str) -> dict:
    used: set = set()
    out: dict = {}
    for row in rows:
        if not row.get("cover_filename"):
            continue
        base = _sanitize(row.get(name_key) or "")
        ext = os.path.splitext(row["cover_filename"])[1] or ".png"
        name, i = base + ext, 2
        while name in used:
            name = f"{base}_{i}{ext}"
            i += 1
        used.add(name)
        out[row["id"]] = name
    return out


def _build_xlsx(
    rows: list, columns: list, title: str, imgname_map: dict, image_bytes_map: dict
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = title
    img_col = None
    for ci, (key, label) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font = Font(bold=True)
        letter = get_column_letter(ci)
        if key == "__image__":
            img_col = ci
            ws.column_dimensions[letter].width = 52
        elif key == "prompt":
            ws.column_dimensions[letter].width = 62
        elif key == "__imgname__":
            ws.column_dimensions[letter].width = 22
        else:
            ws.column_dimensions[letter].width = 16

    for ri, row in enumerate(rows, start=2):
        for ci, (key, _label) in enumerate(columns, start=1):
            if key == "__image__":
                continue
            val = imgname_map.get(row["id"], "") if key == "__imgname__" else row.get(key, "") or ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        placed = False
        if img_col:
            fn = row.get("cover_filename")
            if fn:
                try:
                    xi = _compress_image(image_bytes_map[fn])
                    ratio = (xi.height / xi.width) if xi.width else 0.66
                    xi.width = 360
                    xi.height = int(360 * ratio)
                    ws.add_image(xi, f"{get_column_letter(img_col)}{ri}")
                    ws.row_dimensions[ri].height = xi.height * 0.75
                    placed = True
                except Exception:
                    placed = False
            if not placed:
                ws.cell(row=ri, column=img_col, value="（无图）")
                ws.row_dimensions[ri].height = 56

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_export_zip(rows: list, columns: list, title: str, name_key: str) -> bytes:
    imgname_map = _export_filenames(rows, name_key)
    image_bytes_map = {}
    for row in rows:
        fn = row.get("cover_filename")
        if not fn or fn in image_bytes_map:
            continue
        try:
            image_bytes_map[fn] = storage.get_bytes(fn)
        except Exception:
            continue
    xlsx = _build_xlsx(rows, columns, title, imgname_map, image_bytes_map)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"{title}.xlsx", xlsx)
        for row in rows:
            name = imgname_map.get(row["id"])
            if not name:
                continue
            data = image_bytes_map.get(row["cover_filename"])
            if data:
                zf.writestr(f"原图/{name}", data)
    return buf.getvalue()


def _zip_response(data: bytes, filename: str) -> Response:
    return Response(
        content=data,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"},
    )


@app.get("/api/export/characters")
def export_characters(
    era: Optional[str] = None,
    type: Optional[str] = None,
    gender: Optional[str] = None,
    age: Optional[str] = None,
    genre: Optional[str] = None,
    q: Optional[str] = None,
):
    rows = list_characters(era=era, type=type, gender=gender, age=age, genre=genre, q=q)
    data = _build_export_zip(rows, CHAR_EXPORT_COLS, "角色资产库", "persona")
    return _zip_response(data, "角色资产库.zip")


@app.get("/api/export/scenes")
def export_scenes(
    era: Optional[str] = None,
    scene_type: Optional[str] = None,
    genre: Optional[str] = None,
    mood: Optional[str] = None,
    q: Optional[str] = None,
):
    rows = list_scenes(era=era, scene_type=scene_type, genre=genre, mood=mood, q=q)
    data = _build_export_zip(rows, SCENE_EXPORT_COLS, "场景资产库", "name")
    return _zip_response(data, "场景资产库.zip")


@app.get("/api/health")
def health():
    db_ok = tos_ok = False
    try:
        db_ok = db_health_check()
    except Exception:
        db_ok = False
    try:
        tos_ok = storage.health_check()
    except Exception:
        tos_ok = False
    return {
        "ok": db_ok and tos_ok,
        "db_ok": db_ok,
        "tos_ok": tos_ok,
        "ai_configured": bool(os.getenv("OPENROUTER_API_KEY")),
    }


if __name__ == "__main__":
    import uvicorn
    host = os.getenv("HOST", "127.0.0.1")
    port = int(os.getenv("PORT", "8000"))
    reload = os.getenv("RELOAD", "1") not in ("0", "false", "False")
    uvicorn.run("main:app", host=host, port=port, reload=reload)
